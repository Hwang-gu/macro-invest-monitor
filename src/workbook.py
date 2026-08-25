from __future__ import annotations

from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import START, WORKBOOK_PATH, series_by_key
from .process import load_events

DAILY_MAP = [
    ("kospi", "코스피"),
    ("kosdaq", "코스닥"),
    ("nasdaq", "나스닥"),
    ("sp500", "S&P500"),
    ("bitcoin", "비트코인"),
    ("gold", "금"),
    ("us_3m", "미국국채_단기(3개월)"),
    ("us_10y", "미국국채_장기(10년)"),
    ("us_ffr", "미국_기준금리"),
    ("kr_call", "한국_기준금리"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PRICE_COLS = ["코스피", "코스닥", "나스닥", "S&P500", "금", "비트코인"]
RATE_COLS = ["미국국채_단기(3개월)", "미국국채_장기(10년)", "미국_기준금리", "한국_기준금리"]


def calendar_daily(panel: pd.DataFrame) -> pd.DataFrame:
    """2000-01-01~마지막 시세일을 달력으로 채웁니다. 휴장·주말은 직전 값."""
    end = pd.Timestamp(panel.index.max()).normalize()
    idx = pd.date_range(pd.Timestamp(START), end, freq="D")
    keys = [k for k, _ in DAILY_MAP if k in panel.columns]
    src = panel[keys].copy()
    src.index = pd.to_datetime(src.index).normalize()
    src = src[~src.index.duplicated(keep="last")].sort_index()
    raw = src.reindex(idx)
    filled = raw.copy()
    for col in filled.columns:
        first = filled[col].first_valid_index()
        if first is None:
            continue
        filled.loc[first:, col] = filled.loc[first:, col].ffill()

    events = load_events()
    names: list[str] = []
    for d in idx:
        hit = events.loc[(events["date"] <= d) & (events["end_date"] >= d), "name_ko"]
        names.append("; ".join(hit.tolist()) if len(hit) else "")

    out = pd.DataFrame({"날짜": idx.strftime("%Y-%m-%d")})
    for key, label in DAILY_MAP:
        if key in filled.columns:
            out[label] = filled[key].to_numpy()
            out[f"{label}_원자료있음"] = raw[key].notna().to_numpy()
    out["주말"] = idx.dayofweek >= 5
    out["진행중_이슈"] = names
    out["이슈있음"] = [1 if n else 0 for n in names]
    return out


def issue_table() -> pd.DataFrame:
    events = load_events()
    df = pd.DataFrame({
        "시작": events["date"].dt.strftime("%Y-%m-%d"),
        "종료": events["end_date"].dt.strftime("%Y-%m-%d"),
        "이슈": events["name_ko"],
        "유형": events["category"],
        "강도": events["severity"],
        "설명": events["summary_ko"],
        "id": events["id"],
    })
    return df


def monthly_graph_data(daily: pd.DataFrame) -> pd.DataFrame:
    g = daily.copy()
    g["날짜"] = pd.to_datetime(g["날짜"])
    g = g.set_index("날짜")
    cols = [c for c in PRICE_COLS + RATE_COLS + ["이슈있음"] if c in g.columns]
    m = g[cols].resample("ME").last()
    out = pd.DataFrame({"연월": m.index.strftime("%Y-%m")})
    for col in PRICE_COLS:
        if col not in m:
            continue
        s = m[col]
        first = s.dropna().iloc[0] if s.notna().any() else None
        out[col] = (s / first * 100.0) if first else s
    for col in RATE_COLS:
        if col in m:
            out[col] = m[col]
    out["이슈있음"] = m["이슈있음"] if "이슈있음" in m else 0
    return out


def _write_df(ws: Worksheet, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    for r_i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_i, value in enumerate(row, start=start_col):
            cell = ws.cell(r_i, c_i, value)
            if r_i > start_row and isinstance(value, float):
                cell.number_format = "0.00"


def _paint_header(ws: Worksheet, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _autosize(ws: Worksheet, max_width: int = 24) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        header = str(ws.cell(ws.freeze_panes[0] if False else 1, col).value or "")
        ws.column_dimensions[letter].width = min(max_width, max(12, len(header) + 2))


def _score_frame(mapping: dict | None) -> pd.DataFrame:
    meta = series_by_key()
    if not mapping:
        return pd.DataFrame(columns=["이름", "점수"])
    rows = [{"이름": meta[k].label_ko if k in meta else k, "점수": v} for k, v in mapping.items()]
    return pd.DataFrame(rows).sort_values("점수", ascending=False)


def _forecast_frame(report: dict[str, Any]) -> pd.DataFrame:
    fc = report.get("forecast") or {}
    reasons = " | ".join(fc.get("reasons") or [])
    if not reasons:
        reasons = ((report.get("commentary") or {}).get("display") or "")[:800]
    return pd.DataFrame([
        ("기준일", report.get("asof")),
        ("생성일", report.get("generated")),
        ("시계", fc.get("horizon")),
        ("1단계 자산군 (금 / 비트코인 / 주식)", fc.get("stage1_asset_class")),
        ("2단계 주식이라면 어느 시장", fc.get("stage2_market") or "—"),
        ("3단계 그 시장에서 어느 업종", fc.get("stage3_sector") or "—"),
        ("4단계 코스피라면 대표 종목", fc.get("stage4_stock") or "— (코스피가 아닐 때)"),
        ("종목 코드", fc.get("stage4_ticker") or "—"),
        ("종목 참고", fc.get("stage4_note") or "—"),
        ("근거", reasons),
        ("면책", report.get("disclaimer")),
    ], columns=["항목", "내용"])


def update_workbook(
    panel: pd.DataFrame,
    features: pd.DataFrame,
    report: dict[str, Any],
    channels: list[dict[str, Any]],
    corr: dict[str, Any],
) -> None:
    del features, channels, corr  # 이 장부는 Daily/Issue/Graph/Forecast만 사용
    daily = calendar_daily(panel)
    issues = issue_table()
    monthly = monthly_graph_data(daily)
    forecast = _forecast_frame(report)
    assets = _score_frame(report.get("asset_scores"))
    markets = _score_frame(report.get("market_scores"))
    sectors = _score_frame(report.get("sector_scores"))

    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_d = wb.active
    ws_d.title = "Daily Data"
    ws_d["A1"] = (
        "달력 기준 매일 1행. 주말·휴장은 직전 거래일 값. 비트코인은 2014-09 이전 없음. "
        "금은 선물 근월(현물 대용). 한국 기준금리는 한은 정책금리(BIS 일별, 키 있으면 ECOS). "
    )
    ws_d.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(14, daily.shape[1]))
    ws_d["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    ws_d.row_dimensions[1].height = 40
    _write_df(ws_d, daily, start_row=3)
    _paint_header(ws_d, 3, daily.shape[1])
    ws_d.freeze_panes = "B4"
    last_row = 2 + 1 + len(daily)
    ws_d.auto_filter.ref = f"A3:{get_column_letter(daily.shape[1])}{last_row}"
    for col in range(1, daily.shape[1] + 1):
        ws_d.column_dimensions[get_column_letter(col)].width = 18

    ws_i = wb.create_sheet("Issue")
    _write_df(ws_i, issues)
    _paint_header(ws_i, 1, issues.shape[1])
    ws_i.freeze_panes = "A2"
    ws_i.auto_filter.ref = ws_i.dimensions
    ws_i.column_dimensions["C"].width = 36
    ws_i.column_dimensions["F"].width = 64

    ws_g = wb.create_sheet("Graph")
    ws_g["A1"] = (
        "월말 기준으로 그렸습니다. 주황 막대(이슈있음=1)가 올라온 달이 Issue 구간에 해당합니다. "
        "일별 선 + 이슈 음영은 data/reports/daily_brief.html 또는 streamlit 대시보드가 더 정확합니다."
    )
    ws_g.merge_cells("A1:G1")
    ws_g["A1"].alignment = Alignment(wrap_text=True)
    ws_g.row_dimensions[1].height = 36
    ws_g["A3"] = "이슈 구간"
    _write_df(ws_g, issues[["시작", "종료", "이슈", "유형", "강도"]], start_row=4)
    _paint_header(ws_g, 4, 5)

    start_m = 6 + len(issues)
    ws_g.cell(start_m - 1, 1, "월별 그래프 데이터 (자산은 첫 관측=100)")
    _write_df(ws_g, monthly, start_row=start_m)
    _paint_header(ws_g, start_m, monthly.shape[1])
    n = len(monthly)
    header_row = start_m
    data_end = start_m + n
    cats = Reference(ws_g, min_col=1, min_row=header_row + 1, max_row=data_end)

    price_chart = LineChart()
    price_chart.title = "코스피·코스닥·나스닥·S&P500·금·비트코인 (첫 관측=100)"
    price_chart.y_axis.title = "지수 (시작=100)"
    price_chart.x_axis.title = "연월"
    price_chart.height = 10
    price_chart.width = 22
    price_chart.style = 10
    n_price = sum(1 for c in PRICE_COLS if c in monthly.columns)
    data_ref = Reference(ws_g, min_col=2, max_col=1 + n_price, min_row=header_row, max_row=data_end)
    price_chart.add_data(data_ref, titles_from_data=True)
    price_chart.set_categories(cats)
    price_chart.legend.position = "b"

    issue_col = monthly.columns.get_loc("이슈있음") + 1
    bar = BarChart()
    bar.y_axis.axId = 200
    bar.add_data(Reference(ws_g, min_col=issue_col, min_row=header_row, max_row=data_end), titles_from_data=True)
    bar.set_categories(cats)
    bar.y_axis.title = "이슈 여부"
    price_chart.y_axis.crosses = "min"
    bar.y_axis.crosses = "max"
    price_chart += bar
    ws_g.add_chart(price_chart, "H3")

    rate_chart = LineChart()
    rate_chart.title = "미국 단기·장기 국채, 미국·한국 기준금리"
    rate_chart.y_axis.title = "금리 (%)"
    rate_chart.x_axis.title = "연월"
    rate_chart.height = 8
    rate_chart.width = 22
    rate_start = 2 + n_price
    rate_end = rate_start + sum(1 for c in RATE_COLS if c in monthly.columns) - 1
    rate_chart.add_data(Reference(ws_g, min_col=rate_start, max_col=rate_end, min_row=header_row, max_row=data_end), titles_from_data=True)
    rate_chart.set_categories(cats)
    rate_chart.legend.position = "b"
    ws_g.add_chart(rate_chart, "H22")

    ws_f = wb.create_sheet("Forecast Model")
    _write_df(ws_f, forecast)
    _paint_header(ws_f, 1, 2)
    ws_f.column_dimensions["A"].width = 36
    ws_f.column_dimensions["B"].width = 88
    ws_f["B11"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_f.row_dimensions[11].height = 72

    ws_f["A14"] = "자산군 점수"
    _write_df(ws_f, assets, start_row=15)
    _paint_header(ws_f, 15, 2)
    ws_f["D14"] = "주식 시장 점수"
    _write_df(ws_f, markets, start_row=15, start_col=4)
    _paint_header_range(ws_f, 15, 4, 5)
    ws_f["G14"] = "업종 점수"
    _write_df(ws_f, sectors, start_row=15, start_col=7)
    _paint_header_range(ws_f, 15, 7, 8)

    wb.save(WORKBOOK_PATH)


def _paint_header_range(ws: Worksheet, row: int, c1: int, c2: int) -> None:
    for col in range(c1, c2 + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
